#!/usr/bin/env python3
"""
Export DeepReefMap results to the standardized benthic transect CSV.

Called after each transect is processed by process_transects.sh:

    python3 export_to_csv.py \
        --out_dir /path/to/output/transect_dir \
        --transect_id 8000 \
        --video_name GX010297 \
        --date 30.10.2025 \
        --place "eritrea harat" \
        --transect_quality excellent \
        --length 50m \
        --results_csv results.csv

Metadata lookup order:
  1. transect_gps_complete.xlsx  (by transect_id) — primary source
  2. data/*/3D_sample_list_*.xlsx (by gopro_video_name) — fallback when
     primary GPS fields are absent or contain "TODO"

Output columns match "Résultats transects centrés.csv".
"""

import argparse
import csv
import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

# Canonical benthic category order (template CSV + extra from model output)
BENTHIC_CATEGORIES = [
    "sand",
    "rubble",
    "unknown hard substrate",
    "algae covered substrate",
    "branching bleached",
    "branching dead",
    "branching alive",
    "stylophora alive",
    "pocillopora alive",
    "acropora alive",
    "table acropora alive",
    "table acropora dead",
    "millepora",
    "turbinaria",
    "other coral bleached",
    "other coral dead",
    "other coral alive",
    "massive/meandering alive",
    "massive/meandering dead",
    "massive/meandering bleached",
    "meandering alive",
    "meandering dead",
    "meandering bleached",
    "sea urchin",
    "sea cucumber",
    "anemone",
    "sponge",
    "clam",
    "other animal",
    "trash",
    "seagrass",
    "crown of thorn",
    "dead clam",
]

METADATA_COLS = [
    "transect_id",
    "transect_quality",
    "date",
    "length",
    "place",
    "country",
    "video_filename",
    "transect_start",
    "transect_end",
    "gps_point_name_begin",
    "gps_point_date_begin",
    "gps_point_time_begin",
    "latitude_begin",
    "longitude_begin",
    "gps_accuracy_begin",
    "gps_point_name_end",
    "gps_point_date_end",
    "gps_point_time_end",
    "latitude_end",
    "longitude_end",
    "gps_accuracy_end",
]

ALL_COLS = METADATA_COLS + BENTHIC_CATEGORIES


def _is_missing(value) -> bool:
    """True when a value is absent or a placeholder."""
    if value is None:
        return True
    s = str(value).strip().lower()
    return s in ("", "todo", "n/a", "na", "none")


def load_primary_gps(xlsx_path: Path) -> dict:
    """
    Load transect_gps_complete.xlsx, keyed by integer transect_id.
    Returns a dict of {transect_id: {field: value}}.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    result = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        tid = row[0]
        if tid is None:
            continue
        try:
            tid = int(tid)
        except (ValueError, TypeError):
            continue
        result[tid] = {
            "country": row[1],
            "place": row[2],
            "latitude_begin": row[4],
            "longitude_begin": row[5],
            "gps_accuracy_begin": row[6],
            "latitude_end": row[7],
            "longitude_end": row[8],
            # col 9 is mislabeled gps_accuracy_begin in the xlsx header
            "gps_accuracy_end": row[9],
        }
    return result


def load_secondary_metadata(metadata_dir: Path) -> dict:
    """
    Load all 3D_sample_list_*.xlsx files from metadata_dir, keyed by
    gopro_video_name (the stem of the video file, e.g. "GX010297").
    Returns a dict of {gopro_name: {field: value}}.
    """
    result = {}
    for xlsx in sorted(metadata_dir.glob("3D_sample_list_*.xlsx")):
        wb = openpyxl.load_workbook(xlsx, data_only=True)
        ws = wb.active
        raw_headers = [cell.value for cell in ws[1]]
        # Skip the description row (row 2 has units/descriptions, data starts at row 3)
        # Detect by checking if row 2 looks like a description row
        first_data_row = 2
        sample_row2 = [cell for cell in ws[2]]
        if sample_row2 and isinstance(sample_row2[0].value, str) and sample_row2[0].value not in (None,) and not any(
            c.isdigit() for c in str(sample_row2[0].value or "")
        ):
            first_data_row = 3  # row 2 is descriptions, skip it

        for row in ws.iter_rows(min_row=first_data_row, values_only=True):
            gopro_name = row[raw_headers.index("gopro_video_name")] if "gopro_video_name" in raw_headers else None
            if _is_missing(gopro_name):
                continue
            gopro_name = str(gopro_name).strip()
            # Build a normalised record with the fields we care about
            def get(col):
                try:
                    return row[raw_headers.index(col)]
                except ValueError:
                    return None

            # latitude/longitude: the xlsx uses _start/_end naming
            lat_start = get("latitude_start")
            lon_start = get("longitude_start")
            lat_end   = get("latitude_end")
            lon_end   = get("longitude_end")

            date_val = get("date_iso")
            if hasattr(date_val, "strftime"):
                date_val = date_val.strftime("%d.%m.%Y")

            result[gopro_name] = {
                "country":           get("country"),
                "place":             get("sampling_site_name") or get("region_name"),
                "latitude_begin":    lat_start,
                "longitude_begin":   lon_start,
                "gps_accuracy_begin": None,   # not recorded in this sheet
                "latitude_end":      lat_end,
                "longitude_end":     lon_end,
                "gps_accuracy_end":  None,
                "transect_quality":  get("transect_quality"),
                "length":            get("transect_lentgh"),  # typo in source
                "date":              date_val,
            }
    return result


def merge_metadata(primary: dict, secondary: dict) -> dict:
    """
    Merge two metadata dicts, preferring primary values unless they are missing.
    """
    merged = dict(secondary)
    for key, val in primary.items():
        if not _is_missing(val):
            merged[key] = val
    return merged


def export_result(
    out_dir: Path,
    transect_id: int,
    video_name: str,
    video_filename: str,
    transect_start: str,
    transect_end: str,
    date: str,
    place: str,
    transect_quality: str,
    length: str,
    primary_gps: dict,
    secondary_meta: dict,
    results_csv: Path,
):
    covers_path = out_dir / "percentage_covers.json"
    if not covers_path.exists():
        print(f"ERROR: {covers_path} not found", file=sys.stderr)
        sys.exit(1)

    with open(covers_path) as f:
        covers = json.load(f)

    # --- Merge metadata from both sources ------------------------------------
    p = primary_gps.get(transect_id, {})
    s = secondary_meta.get(video_name, {})
    meta = merge_metadata(p, s)

    # Fields that come from the input CSV take lowest priority
    # (only used if both xlsx sources are missing the value)
    resolved_place         = meta.get("place") or place
    resolved_country       = meta.get("country") or ""
    resolved_quality       = meta.get("transect_quality") or transect_quality
    resolved_length        = meta.get("length") or length
    # Date: prefer secondary (it's a real datetime), then what was passed in
    resolved_date          = meta.get("date") or date

    def gps(field):
        v = meta.get(field)
        return "" if _is_missing(v) else v

    row = {
        "transect_id":         transect_id,
        "transect_quality":    resolved_quality,
        "date":                resolved_date,
        "length":              resolved_length,
        "place":               resolved_place,
        "country":             resolved_country,
        "video_filename":      video_filename,
        "transect_start":      transect_start,
        "transect_end":        transect_end,
        "gps_point_name_begin": "",
        "gps_point_date_begin": "",
        "gps_point_time_begin": "",
        "latitude_begin":      gps("latitude_begin"),
        "longitude_begin":     gps("longitude_begin"),
        "gps_accuracy_begin":  gps("gps_accuracy_begin"),
        "gps_point_name_end":  "",
        "gps_point_date_end":  "",
        "gps_point_time_end":  "",
        "latitude_end":        gps("latitude_end"),
        "longitude_end":       gps("longitude_end"),
        "gps_accuracy_end":    gps("gps_accuracy_end"),
    }

    for cat in BENTHIC_CATEGORIES:
        row[cat] = covers.get(cat, 0)

    write_header = not results_csv.exists() or results_csv.stat().st_size == 0
    with open(results_csv, "a", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=ALL_COLS)
        if write_header:
            writer.writeheader()
        writer.writerow(row)

    print(f"  Exported transect {transect_id} ({video_name}) → {results_csv}")


def main():
    parser = argparse.ArgumentParser(
        description="Export DeepReefMap percentage_covers.json to the results CSV"
    )
    parser.add_argument(
        "--out_dir", required=True,
        help="Output directory containing percentage_covers.json",
    )
    parser.add_argument("--transect_id", required=True, type=int)
    parser.add_argument(
        "--video_name", default="",
        help="GoPro video stem (e.g. GX010297) used to look up secondary metadata",
    )
    parser.add_argument("--video_filename", default="", help="Full relative video path from the input CSV")
    parser.add_argument("--transect_start", default="", help="Transect start timestamp (MM:SS)")
    parser.add_argument("--transect_end",   default="", help="Transect end timestamp (MM:SS)")
    parser.add_argument("--date", required=True, help="Transect date (fallback)")
    parser.add_argument("--place", default="", help="Site name (fallback)")
    parser.add_argument("--transect_quality", default="")
    parser.add_argument("--length", default="")
    parser.add_argument(
        "--gps_xlsx", default=None,
        help="Path to transect_gps_complete.xlsx",
    )
    parser.add_argument(
        "--metadata_dir", default=None,
        help="Directory containing 3D_sample_list_*.xlsx files "
             "(default: <project_root>/data/TRSC_er_102025_metadata)",
    )
    parser.add_argument(
        "--results_csv", default=None,
        help="Output CSV path (default: <project_root>/results.csv)",
    )
    args = parser.parse_args()

    project_root = Path(__file__).parent.parent
    gps_xlsx    = Path(args.gps_xlsx)    if args.gps_xlsx    else project_root / "data" / "transect_gps_complete.xlsx"
    meta_dir    = Path(args.metadata_dir) if args.metadata_dir else project_root / "data" / "TRSC_er_102025_metadata"
    results_csv = Path(args.results_csv) if args.results_csv else project_root / "results.csv"

    primary_gps = {}
    if gps_xlsx.exists():
        primary_gps = load_primary_gps(gps_xlsx)
    else:
        print(f"WARNING: GPS file not found: {gps_xlsx}", file=sys.stderr)

    secondary_meta = {}
    if meta_dir.exists():
        secondary_meta = load_secondary_metadata(meta_dir)
        if secondary_meta:
            print(f"  Loaded secondary metadata for {len(secondary_meta)} videos from {meta_dir.name}/")
    else:
        print(f"WARNING: metadata dir not found: {meta_dir}", file=sys.stderr)

    export_result(
        out_dir=Path(args.out_dir),
        transect_id=args.transect_id,
        video_name=args.video_name,
        video_filename=args.video_filename,
        transect_start=args.transect_start,
        transect_end=args.transect_end,
        date=args.date,
        place=args.place,
        transect_quality=args.transect_quality,
        length=args.length,
        primary_gps=primary_gps,
        secondary_meta=secondary_meta,
        results_csv=results_csv,
    )


if __name__ == "__main__":
    main()
